"""
Tests for reporting_utils.py - shared Excel reporting utilities
"""

import pytest
from openpyxl import Workbook
from openpyxl.chart import BarChart
from openpyxl.chart.label import DataLabelList

from econ_capital.reporting_utils import apply_clean_style, autofit_columns
from openpyxl.chart import Reference


@pytest.fixture
def sample_chart_fixture():
    """Create a worksheet with sample data and a bar chart (11 points)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "TestReport"

    # Sample data - 11 categories (matches num_points default)
    categories = [
        "Jan",
        "Feb",
        "Mar",
        "Apr",
        "May",
        "Jun",
        "Jul",
        "Aug",
        "Sep",
        "Oct",
        "Nov",
    ]
    values = [12.5, 8.7, 15.2, 9.8, 22.1, 18.4, 14.7, 11.3, 19.6, 16.8, 13.9]

    # Headers and data
    ws["A1"] = "Month"
    ws["B1"] = "Loss (£M)"

    for i, (cat, val) in enumerate(zip(categories, values), start=2):
        ws[f"A{i}"] = cat
        ws[f"B{i}"] = val

    # Create basic column (bar) chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Monthly Losses"
    chart.y_axis.title = "Loss Amount"
    chart.x_axis.title = "Month"

    data = Reference(ws, min_col=2, min_row=1, max_row=12)
    cats = Reference(ws, min_col=1, min_row=2, max_row=12)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, "D2")

    return ws, chart


class TestApplyCleanStyle:
    def test_applies_y_axis_title(self, sample_chart_fixture):
        ws, chart = sample_chart_fixture
        apply_clean_style(chart, "Total Loss (£M)")
        assert chart.y_axis.title.tx.rich.p[0].r[0].t == "Total Loss (£M)"

    def test_removes_legend(self, sample_chart_fixture):
        ws, chart = sample_chart_fixture
        apply_clean_style(chart, "Loss")
        assert chart.legend is None

    def test_sets_vary_colors(self, sample_chart_fixture):
        ws, chart = sample_chart_fixture
        apply_clean_style(chart, "Loss")
        assert chart.varyColors is True

    def test_applies_number_format_millions(self, sample_chart_fixture):
        ws, chart = sample_chart_fixture
        apply_clean_style(chart, "Loss (£M)")
        assert chart.y_axis.number_format.formatCode == '"£"#,##0,,"M"'

    def test_sets_manual_layout(self, sample_chart_fixture):
        ws, chart = sample_chart_fixture
        apply_clean_style(chart, "Loss")
        assert chart.layout is not None
        assert chart.layout.manualLayout is not None
        ml = chart.layout.manualLayout
        assert ml.x == 0.1
        assert ml.y == 0.05
        assert ml.w == 0.85
        assert ml.h == 0.70

    def test_removes_y_gridlines(self, sample_chart_fixture):
        ws, chart = sample_chart_fixture
        apply_clean_style(chart, "Loss")
        assert chart.y_axis.majorGridlines is None

    def test_applies_x_axis_label_rotation(self, sample_chart_fixture):
        ws, chart = sample_chart_fixture
        apply_clean_style(chart, "Loss")
        assert chart.x_axis.textRotation == -45

    def test_applies_color_palette_to_first_series(self, sample_chart_fixture):
        ws, chart = sample_chart_fixture
        apply_clean_style(chart, "Loss", num_points=11)

        series = chart.series[0]
        assert len(series.dPt) == 11

        expected_colors = [
            "F06292",
            "E67E22",
            "BCAE3E",
            "7CB342",
            "27AE60",
            "26A69A",
            "2980B9",
            "5D9CEC",
            "BA68C8",
            "9575CD",
            "F06292",
        ]

        for i, dp in enumerate(series.dPt):
            assert dp.graphicalProperties.solidFill.srgbClr == expected_colors[i]
            assert dp.invertIfNegative is False
            assert dp.graphicalProperties.ln is None  # no outline

    def test_adds_data_labels_when_none_exist(self, sample_chart_fixture):
        ws, chart = sample_chart_fixture
        apply_clean_style(chart, "Loss")

        series = chart.series[0]
        assert series.dLbls is not None
        assert series.dLbls.showVal is True
        assert series.dLbls.showCatName is False
        assert series.dLbls.showSerName is False
        assert series.dLbls.showLegendKey is False
        assert series.dLbls.position == "outEnd"
        assert series.dLbls.showLeaderLines is True
        assert series.dLbls.numFmt == '"£"#,##0,,"M"'

    def test_respects_existing_data_labels(self, sample_chart_fixture):
        ws, chart = sample_chart_fixture
        # Pre-set custom label settings
        if not chart.series:
            pytest.skip("No series created in fixture")
        chart.series[0].dLbls = DataLabelList()
        chart.series[0].dLbls.showVal = False  # different from default

        apply_clean_style(chart, "Loss")

        # Should NOT override existing configuration
        assert chart.series[0].dLbls.showVal is False


class TestAutofitColumns:
    @pytest.fixture
    def worksheet_with_various_lengths(self):
        wb = Workbook()
        ws = wb.active

        # Column A - long header + mixed content
        ws["A1"] = (
            "Very Very Very Very Very Long Column Header That Should Be Truncated"
        )
        ws["A2"] = "Short"
        ws["A3"] = None
        ws["A4"] = "Medium length text here"

        # Column B - longer description + very long word
        ws["B1"] = "Category"
        ws["B2"] = (
            "This is a much longer description that needs wrapping and should be visible"
        )
        ws["B3"] = "VeryLongWordWithoutSpacesThatMightBeProblematic"

        # Column C - extremely long content to test cap
        ws["C1"] = "x" * 200

        return ws

    def test_autofits_columns_with_reasonable_widths(
        self, worksheet_with_various_lengths
    ):
        ws = worksheet_with_various_lengths
        autofit_columns(ws)

        assert 30 <= ws.column_dimensions["A"].width <= 50
        assert 30 <= ws.column_dimensions["B"].width <= 50

    def test_sets_wrap_text_and_vertical_alignment(
        self, worksheet_with_various_lengths
    ):
        ws = worksheet_with_various_lengths
        autofit_columns(ws)

        # Correct iteration over columns (ws.columns returns tuples of Cell objects)
        for col_cells in ws.columns:
            for cell in col_cells:
                if cell.value is not None:
                    assert cell.alignment.wrap_text is True
                    assert cell.alignment.vertical == "top"

    def test_limits_max_width_to_50(self, worksheet_with_various_lengths):
        ws = worksheet_with_various_lengths
        autofit_columns(ws)

        assert ws.column_dimensions["C"].width == 50
