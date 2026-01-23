"""
Regulatory-grade reporting for Market Risk Economic Capital.
"""

from __future__ import annotations
from datetime import datetime
from pathlib import Path
from typing import Dict, Any
import pandas as pd
import numpy as np

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

from .engine import MarketRiskEconomicCapital
from .covariance import ewma_cov

from econ_capital.reporting_utils import apply_clean_style, autofit_columns


class MarketRiskReporter:
    def __init__(
        self,
        engine: MarketRiskEconomicCapital,
        results: Dict[str, Any],
        config: Dict[str, Any],
        output_dir: str | Path = "reports",
    ):
        self.engine = engine
        self.results = results
        self.config = config
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.filename = (
            self.output_dir
            / f"MarketRisk_EC_Report_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        )

        self.colors = self.config.get("reporting_style", {}).get(
            "colors", self._default_colors()
        )
        # Ensure fallback for required keys
        default_colors = self._default_colors()
        if "gold" not in self.colors:
            self.colors["gold"] = default_colors["gold"]
        if "header" not in self.colors:
            self.colors["header"] = default_colors["header"]

    def _default_colors(self):
        """Standard hardcoded colors for reporting style used as a fallback."""
        return {
            "header": "1F4E78",  # Dark Blue/Grey for headers
            "table_header": "DDEBF7",  # Light Blue/Grey fill
            "gold": "FFD700",  # Gold for top ranks
            "white": "FFFFFF",
        }

    def generate_full_report(self) -> Path:
        wb = Workbook()
        wb.remove(wb.active)

        self._create_cover_sheet(wb)
        self._create_summary_sheet(wb)
        self._create_waterfall_sheet(wb)
        self._create_covar_sheet(wb)
        self._create_risk_factor_analysis_sheet(wb)
        self._create_stress_testing_sheet(wb)
        self._create_historical_var_sheet(wb)
        self._create_methodology_sheet(wb)

        wb.save(self.filename)
        print(f"Report generated: {self.filename}")
        return self.filename

    @property
    def n_paths(self) -> int:
        return self.engine.config.get("n_paths", 500_000)  # fallback to default

    @property
    def cov_method(self) -> str:
        return self.engine.config.get("cov_method", "EWMA").upper()

    def _create_cover_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Cover", 0)
        lines = [
            "MARKET RISK",
            "ECONOMIC CAPITAL",
            "REPORT",
            "",
            f"Run Date: {datetime.now():%Y-%m-%d %H:%M}",
            f"10D VaR (99.9%): £{self.results['var_10d_999']:,.0f}",
            f"10D ES (99.9%): £{self.results['es_10d_999']:,.0f}",
        ]
        for i, line in enumerate(lines, 2):
            cell = ws[f"A{i}"]
            cell.value = line
            cell.font = Font(
                size=24 if i <= 4 else 14, bold=True, color=self.colors["header"]
            )
            cell.alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:F4")
        autofit_columns(ws)

    def _create_summary_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Summary", 1)
        ws["A1"] = "Executive Summary"
        ws["A1"].font = Font(size=18, bold=True, color=self.colors["header"])

        data = [
            ("Run Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("10D VaR (99.9%)", f"£{self.results['var_10d_999']:,.0f}"),
            ("10D ES (99.9%)", f"£{self.results['es_10d_999']:,.0f}"),
            ("1Y VaR (99.9%)", f"£{self.results['var_1y_999']:,.0f}"),
            ("1Y ES (99.9%)", f"£{self.results['es_1y_999']:,.0f}"),
            ("Number of Paths", self.n_paths),
            ("Covariance Method", self.cov_method),
        ]
        for i, (k, v) in enumerate(data, 3):
            ws[f"A{i}"] = k
            cell = ws[f"B{i}"]
            cell.value = v
            if k == "Number of Paths":
                cell.number_format = "#,##0"
            ws[f"A{i}"].font = Font(bold=True)
        autofit_columns(ws)

    def _create_waterfall_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Waterfall", 3)
        ws["A1"] = "Marginal Capital Contributions (Post-Diversification)"
        ws["A1"].font = Font(size=16, bold=True, color=self.colors["header"])

        # Prepare Data with Headers
        headers = ["Position", "Capital Contribution (£)"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(2, c, h)
            cell.fill = PatternFill("solid", self.colors["header"])
            cell.font = Font(color="FFFFFF", bold=True)

        baseline = self.results.get("baseline_capital", 0)
        breakdown = self.results["capital_breakdown"]

        # Starting data at row 3
        ws.cell(3, 1, "Baseline")
        ws.cell(3, 2, baseline).number_format = "£#,##0"

        current_row = 4
        for i, (pos, uplift) in enumerate(breakdown.head(10).items(), 1):
            name_cell = ws.cell(current_row, 1, pos)
            val_cell = ws.cell(current_row, 2, uplift)
            val_cell.number_format = "£#,##0"

            # Apply gold highlight to top 3 contributors
            if i <= 3:
                gold_fill = PatternFill("solid", fgColor=self.colors["gold"])
                name_cell.fill = gold_fill
                val_cell.fill = gold_fill

            current_row += 1

        table_range = f"A2:B{current_row - 1}"
        tab = Table(displayName="WaterfallTable", ref=table_range)
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True
        )
        ws.add_table(tab)

        # Create chart
        chart = BarChart()
        chart.title = "Top Capital Impacts"
        chart.height = 14
        chart.width = 24
        data_ref = Reference(ws, min_col=2, min_row=3, max_row=current_row - 1)
        cats = Reference(ws, min_col=1, min_row=3, max_row=current_row - 1)
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(cats)

        # Calculate number of points to plot
        num_plot_points = 1 + len(breakdown.head(10))

        apply_clean_style(
            chart,
            "Capital Contribution (£)",
            num_points=num_plot_points,
        )

        ws.add_chart(chart, "D2")
        current_row += 2
        ws.cell(
            current_row,
            1,
            "Note: Individual contributions have been scaled to align with the diversified 1Y Expected Loss.",
        )
        ws.cell(current_row, 1).font = Font(italic=True)

        autofit_columns(ws)

    def _create_covar_sheet(self, wb: Workbook):
        """Display CoVaR metrics showing systemic risk contributions."""
        ws = wb.create_sheet("CoVaR Systemic Risk", 4)

        ws["A1"] = "Conditional Value-at-Risk (CoVaR) - Systemic Risk Metrics"
        ws["A1"].font = Font(size=16, bold=True, color=self.colors["header"])

        # Check if CoVaR was computed
        covar_metrics = self.results.get("covar_metrics", {})

        if not covar_metrics:
            ws["A3"] = "CoVaR analysis not available (set compute_covar=True in config)"
            ws["A3"].font = Font(italic=True, color="FF0000")
            return

        # Note
        ws["A2"] = (
            "CoVaR measures the portfolio VaR conditional on a specific position being stressed. "
            "ΔCoVaR quantifies each position's systemic contribution to portfolio risk."
        )
        ws["A2"].font = Font(size=10, italic=True)
        ws.merge_cells("A2:F2")
        ws["A2"].alignment = Alignment(wrap_text=True)

        # Headers
        headers = [
            "Position",
            "CoVaR (£)",
            "ΔCoVaR (£)",
            "Systemic Contrib. (%)",
            "Risk Classification",
        ]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(4, c, h)
            cell.fill = PatternFill("solid", self.colors["header"])
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Convert to DataFrame for easier handling
        covar_df = pd.DataFrame(covar_metrics).T
        covar_df = covar_df.sort_values("systemic_contribution_pct", ascending=False)

        # Write data with conditional formatting
        for i, (pos_name, row) in enumerate(covar_df.iterrows(), 5):
            ws.cell(i, 1, pos_name)

            # CoVaR
            ws.cell(i, 2, float(row["covar"])).number_format = "£#,##0"

            # ΔCoVaR
            delta_covar = float(row["delta_covar"])
            cell = ws.cell(i, 3, delta_covar)
            cell.number_format = "£#,##0"

            # Conditional color: red if large positive (risk amplifier)
            if delta_covar > 0:
                cell.fill = PatternFill("solid", fgColor="FFC7CE")  # Light red
                cell.font = Font(color="9C0006", bold=True)

            # Systemic %
            sys_pct = float(row["systemic_contribution_pct"])
            ws.cell(i, 4, sys_pct / 100).number_format = "0.00%"

            # Risk Classification
            if sys_pct > 10:
                classification = "High Systemic"
                color = "FFC7CE"
            elif sys_pct > 5:
                classification = "Moderate Systemic"
                color = "FFEB9C"
            else:
                classification = "Low Systemic"
                color = "C6EFCE"

            cell = ws.cell(i, 5, classification)
            cell.fill = PatternFill("solid", fgColor=color)
            cell.alignment = Alignment(horizontal="center")

        # Table
        last_row = len(covar_df) + 4
        tab = Table(displayName="CoVaRTable", ref=f"A4:E{last_row}")
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showRowStripes=True
        )
        ws.add_table(tab)

        # Chart: ΔCoVaR contributions
        chart = BarChart()
        chart.type = "col"
        chart.title = "Systemic Risk Contributions (ΔCoVaR)"
        chart.x_axis.title = "Position"
        chart.height = 14
        chart.width = 20

        data_ref = Reference(ws, min_col=3, min_row=5, max_row=last_row)
        cats = Reference(ws, min_col=1, min_row=5, max_row=last_row)
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(cats)

        apply_clean_style(chart, "ΔCoVaR (£)")
        ws.add_chart(chart, "G2")

        # Footnote
        footnote_row = last_row + 3
        ws.cell(footnote_row, 1, "Interpretation:")
        ws.cell(footnote_row, 1).font = Font(bold=True, color=self.colors["header"])

        footnote_text = (
            "• CoVaR = Portfolio VaR conditional on position at its own VaR level\n"
            "• ΔCoVaR = CoVaR - Baseline Portfolio VaR (incremental systemic risk)\n"
            "• Positive ΔCoVaR indicates position amplifies portfolio risk when stressed\n"
            "• High systemic contributors (>10%) warrant enhanced monitoring and limits"
        )

        ws.cell(footnote_row, 2, footnote_text)
        ws.cell(footnote_row, 2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(footnote_row, 2).font = Font(size=9, italic=True)
        ws.merge_cells(f"B{footnote_row}:F{footnote_row + 3}")

        autofit_columns(ws)

    def _create_risk_factor_analysis_sheet(self, wb: Workbook):
        """Detailed factor-level risk decomposition."""
        ws = wb.create_sheet("Risk Factor Analysis", 5)

        ws["A1"] = "Risk Factor Variance-Covariance Analysis"
        ws["A1"].font = Font(size=16, bold=True, color=self.colors["header"])

        # Get factor stats from engine
        risk_factors = self.engine.risk_factors
        factor_names = list(risk_factors.columns)

        # Compute recent volatilities (EWMA)

        cov_matrix = ewma_cov(risk_factors, lamb=0.97)
        vols = np.sqrt(np.diag(cov_matrix))

        # Annualized volatilities (252 trading days)
        vols_annual = vols * np.sqrt(252)

        # Headers
        headers = [
            "Risk Factor",
            "Daily Vol (%)",
            "Annual Vol (%)",
            "Recent Return (Ann. %)",
            "Vol Regime",
            "Contribution to Portfolio Vol",
        ]

        for c, h in enumerate(headers, 1):
            cell = ws.cell(3, c, h)
            cell.fill = PatternFill("solid", self.colors["header"])
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Portfolio volatility using equal weights for demonstration (simplified)
        weights = np.ones(len(factor_names)) / len(factor_names)
        port_var = weights @ cov_matrix.values @ weights
        port_vol = np.sqrt(port_var)

        # Write data
        start_row = 4

        # Ensure vols and vols_annual are arrays/lists with correct length
        if len(vols) != len(factor_names):
            raise ValueError(
                f"vols length {len(vols)} does not match {len(factor_names)} factors"
            )
        if len(vols_annual) != len(factor_names):
            raise ValueError("vols_annual length mismatch")

        # Most recent returns as Series indexed by factor names
        recent_returns_series = risk_factors.iloc[-1]  # last row (most recent day)

        for row, factor in enumerate(factor_names, start=start_row):
            idx = row - start_row

            ws.cell(row, 1, factor)

            # Daily volatility
            ws.cell(row, 2, vols[idx]).number_format = "0.00%"

            # Annual volatility
            annual_vol = vols_annual[idx]
            cell = ws.cell(row, 3, annual_vol)
            cell.number_format = "0.00%"

            # Color code by volatility level
            if annual_vol > 30:
                cell.fill = PatternFill("solid", fgColor="FFC7CE")  # High vol - red
            elif annual_vol > 20:
                cell.fill = PatternFill("solid", fgColor="FFEB9C")  # Med vol - yellow

            recent_ret = recent_returns_series[factor] * 100
            ws.cell(row, 4, recent_ret).number_format = "0.00%"

            # Volatility regime
            recent_vol = risk_factors[factor].tail(20).std() * np.sqrt(252)
            long_vol = risk_factors[factor].tail(60).std() * np.sqrt(252)

            if recent_vol > 1.5 * long_vol:
                regime = "High Vol"
                regime_color = "FFC7CE"
            elif recent_vol < 0.7 * long_vol:
                regime = "Low Vol"
                regime_color = "C6EFCE"
            else:
                regime = "Normal"
                regime_color = "FFFFFF"

            cell = ws.cell(row, 5, regime)
            cell.fill = PatternFill("solid", fgColor=regime_color)
            cell.alignment = Alignment(horizontal="center")

            # Contribution to portfolio vol (marginal contribution)
            marginal_contrib = (cov_matrix.iloc[idx, :].values @ weights) / port_vol
            ws.cell(row, 6, marginal_contrib * 100).number_format = "0.00%"

        # Table
        last_row = len(factor_names) + 3
        tab = Table(displayName="FactorAnalysisTable", ref=f"A3:F{last_row}")
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True
        )
        ws.add_table(tab)

        # Correlation heatmap (simplified - just list top correlations)
        ws.cell(last_row + 3, 1, "Top Correlated Factor Pairs").font = Font(
            size=14, bold=True, color=self.colors["header"]
        )

        # Extract upper triangle correlations
        corr_matrix = risk_factors.corr()
        corr_pairs = []
        for i in range(len(factor_names)):
            for j in range(i + 1, len(factor_names)):
                corr_pairs.append(
                    (factor_names[i], factor_names[j], corr_matrix.iloc[i, j])
                )

        # Sort by absolute correlation
        corr_pairs.sort(key=lambda x: abs(x[2]), reverse=True)

        # Headers for correlation pairs
        corr_start = last_row + 4
        ws.cell(corr_start, 1, "Factor 1").font = Font(bold=True)
        ws.cell(corr_start, 2, "Factor 2").font = Font(bold=True)
        ws.cell(corr_start, 3, "Correlation").font = Font(bold=True)

        for c in range(1, 4):
            ws.cell(corr_start, c).fill = PatternFill(
                "solid", self.colors["table_header"]
            )

        # Top 10 correlations
        for idx, (f1, f2, corr) in enumerate(corr_pairs[:10], corr_start + 1):
            ws.cell(idx, 1, f1)
            ws.cell(idx, 2, f2)
            cell = ws.cell(idx, 3, corr)
            cell.number_format = "0.00"

            # Color code
            if abs(corr) > 0.7:
                cell.fill = PatternFill("solid", fgColor="FFC7CE")  # Strong correlation
            elif abs(corr) > 0.4:
                cell.fill = PatternFill("solid", fgColor="FFEB9C")

        autofit_columns(ws)

    def _create_stress_testing_sheet(self, wb: Workbook):
        """Display stress test results and scenario analysis."""
        ws = wb.create_sheet("Stress Testing", 6)

        ws["A1"] = "Market Risk Stress Testing"
        ws["A1"].font = Font(size=16, bold=True, color=self.colors["header"])

        # Check if stress testing was enabled
        stressed_var = self.results.get("stressed_var_1y_999")
        stressed_es = self.results.get("stressed_es_1y_999")

        if stressed_var is None or stressed_es is None:
            ws["A3"] = "Stress testing not enabled (set stress_enabled=True in config)"
            ws["A3"].font = Font(italic=True, color="FF0000")
            return

        # Summary comparison
        ws["A3"] = "Baseline vs. Stressed Metrics Comparison"
        ws["A3"].font = Font(size=14, bold=True, color=self.colors["header"])

        headers = [
            "Metric",
            "Baseline (£)",
            "Stressed (£)",
            "Increase (£)",
            "Increase (%)",
        ]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(4, c, h)
            cell.fill = PatternFill("solid", self.colors["header"])
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Baseline metrics
        baseline_var = self.results["var_1y_999"]
        baseline_es = self.results["es_1y_999"]

        # VaR comparison
        ws.cell(5, 1, "1Y VaR (99.9%)")
        ws.cell(5, 2, baseline_var).number_format = "£#,##0"
        ws.cell(5, 3, stressed_var).number_format = "£#,##0"

        var_increase = stressed_var - baseline_var
        ws.cell(5, 4, var_increase).number_format = "£#,##0"
        ws.cell(
            5, 5, var_increase / baseline_var if baseline_var != 0 else 0
        ).number_format = "0.00%"

        # ES comparison
        ws.cell(6, 1, "1Y ES (99.9%)")
        ws.cell(6, 2, baseline_es).number_format = "£#,##0"
        ws.cell(6, 3, stressed_es).number_format = "£#,##0"

        es_increase = stressed_es - baseline_es
        ws.cell(6, 4, es_increase).number_format = "£#,##0"
        ws.cell(
            6, 5, es_increase / baseline_es if baseline_es != 0 else 0
        ).number_format = "0.00%"

        # Stress scenario details
        stress_shocks = self.config.get("stress_shocks", {})

        if stress_shocks:
            ws.cell(9, 1, "Applied Stress Shocks").font = Font(
                size=14, bold=True, color=self.colors["header"]
            )

            shock_headers = ["Risk Factor", "Shock (%)", "Interpretation"]
            for c, h in enumerate(shock_headers, 1):
                cell = ws.cell(10, c, h)
                cell.fill = PatternFill("solid", self.colors["table_header"])
                cell.font = Font(bold=True)

            for i, (factor, shock) in enumerate(stress_shocks.items(), 11):
                ws.cell(i, 1, factor)
                ws.cell(i, 2, shock * 100).number_format = "0.00%"

                # Interpretation
                if shock < -0.3:
                    interp = "Severe Market Crash"
                elif shock < -0.1:
                    interp = "Significant Downturn"
                elif shock > 0.1:
                    interp = "Major Upward Move"
                else:
                    interp = "Moderate Stress"

                ws.cell(i, 3, interp)

        # Chart: Baseline vs Stressed
        chart = BarChart()
        chart.type = "col"
        chart.title = "Baseline vs. Stressed Capital Requirements"
        chart.height = 12
        chart.width = 18

        # Data for both VaR and ES
        data_ref = Reference(ws, min_col=2, min_row=4, max_row=6, max_col=3)
        cats = Reference(ws, min_col=1, min_row=5, max_row=6)

        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)

        # Series colors
        if chart.series:
            chart.series[
                0
            ].graphicalProperties.solidFill = "4472C4"  # Blue for baseline
            chart.series[1].graphicalProperties.solidFill = "C5504B"  # Red for stressed

        apply_clean_style(chart, "Economic Capital (£)")
        ws.add_chart(chart, "G2")

        # Regulatory note
        note_row = 16
        ws.cell(note_row, 1, "Regulatory Context:")
        ws.cell(note_row, 1).font = Font(bold=True, color=self.colors["header"])

        note_text = (
            "Stress testing is a key component of the Internal Capital Adequacy Assessment Process (ICAAP). "
            "The stressed scenarios reflect severe but plausible market conditions aligned with regulatory "
            "expectations (e.g., PRA SS31/15, EBA stress testing guidelines). Management should ensure that "
            "capital buffers exceed stressed requirements under all adverse scenarios."
        )

        ws.cell(note_row, 2, note_text)
        ws.cell(note_row, 2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(note_row, 2).font = Font(size=9, italic=True)
        ws.merge_cells(f"B{note_row}:F{note_row + 3}")

        autofit_columns(ws)

    def _create_historical_var_sheet(self, wb: Workbook):
        """Compare historical simulation vs. parametric VaR."""
        ws = wb.create_sheet("Historical vs Parametric", 7)

        ws["A1"] = "Historical Simulation vs. Parametric VaR"
        ws["A1"].font = Font(size=16, bold=True, color=self.colors["header"])

        # Check if historical VaR was computed
        hist_var = self.results.get("historical_var_1y_999")
        hist_es = self.results.get("historical_es_1y_999")

        if hist_var is None:
            ws["A3"] = (
                "Historical VaR not enabled (set use_historical_var=True in config)"
            )
            ws["A3"].font = Font(italic=True, color="FF0000")
            return

        # Comparison table
        headers = [
            "Method",
            "1Y VaR (99.9%)",
            "1Y ES (99.9%)",
            "Difference from Parametric",
        ]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(3, c, h)
            cell.fill = PatternFill("solid", self.colors["header"])
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        param_var = self.results["var_1y_999"]
        param_es = self.results["es_1y_999"]

        # Parametric row
        ws.cell(4, 1, "Parametric (t-Copula)")
        ws.cell(4, 2, param_var).number_format = "£#,##0"
        ws.cell(4, 3, param_es).number_format = "£#,##0"
        ws.cell(4, 4, "Baseline")

        # Historical row
        ws.cell(5, 1, "Historical Simulation")
        ws.cell(5, 2, hist_var).number_format = "£#,##0"
        ws.cell(5, 3, hist_es).number_format = "£#,##0"

        diff_pct = ((hist_var - param_var) / param_var * 100) if param_var != 0 else 0
        cell = ws.cell(5, 4, diff_pct / 100)
        cell.number_format = "0.00%"

        # Color code difference
        if abs(diff_pct) > 20:
            cell.fill = PatternFill("solid", fgColor="FFC7CE")  # Large divergence - red
        elif abs(diff_pct) > 10:
            cell.fill = PatternFill("solid", fgColor="FFEB9C")  # Moderate - yellow

        # Interpretation
        ws.cell(7, 1, "Interpretation:")
        ws.cell(7, 1).font = Font(bold=True, color=self.colors["header"])

        if abs(diff_pct) < 10:
            interpretation = "Parametric and Historical VaR are closely aligned, indicating stable market conditions."
        elif hist_var > param_var:
            interpretation = "Historical VaR exceeds Parametric VaR, suggesting actual market tail events are more severe than assumed by the model. Consider increasing model parameters."
        else:
            interpretation = "Parametric VaR exceeds Historical VaR, indicating the model may be conservative or recent history was benign."

        ws.cell(7, 2, interpretation)
        ws.cell(7, 2).alignment = Alignment(wrap_text=True)
        ws.cell(7, 2).font = Font(size=10, italic=True)
        ws.merge_cells("B7:D8")

        autofit_columns(ws)

    def _create_methodology_sheet(self, wb: Workbook):
        """Technical documentation of methods used."""
        ws = wb.create_sheet("Methodology", 8)

        ws["A1"] = "Market Risk Methodology & Model Documentation"
        ws["A1"].font = Font(size=16, bold=True, color=self.colors["header"])

        # Model configuration
        config_section = [
            ("Configuration", ""),
            ("Simulation Paths", f"{self.n_paths:,}"),
            ("Horizon", f"{self.engine.horizon_days} days"),
            ("Confidence Level", f"{self.engine.var_q * 100:.1f}%"),
            ("Covariance Method", self.cov_method),
            (
                "Degrees of Freedom (t-dist)",
                f"{self.engine.config.get('df_t', 3.0):.1f}",
            ),
            (
                "Full Revaluation Enabled",
                str(self.results.get("used_full_revaluation", False)),
            ),
            ("", ""),
            ("Data Window", ""),
            ("Start Date", self.config.get("start_date", "2021-01-01")),
            ("End Date", self.config.get("end_date", "2026-01-01")),
            ("Observations", len(self.engine.risk_factors)),
        ]

        ws.cell(3, 1, "Model Configuration").font = Font(
            size=14, bold=True, color=self.colors["header"]
        )

        for i, (param, value) in enumerate(config_section, 4):
            ws.cell(i, 1, param).font = Font(bold=True)
            ws.cell(i, 2, str(value))

        # Methodology description
        method_row = len(config_section) + 6

        ws.cell(method_row, 1, "Methodology Overview").font = Font(
            size=14, bold=True, color=self.colors["header"]
        )

        methodology_text = """
    1. COVARIANCE ESTIMATION
    - EWMA: Exponentially weighted moving average with λ=0.97 (RiskMetrics standard)
    - GARCH(1,1): Univariate volatility models with sample correlation matrix
    - Sample: Unbiased sample covariance (minimum 252 observations recommended)

    2. SHOCK GENERATION
    - Multivariate Student-t distribution with configurable degrees of freedom (ν)
    - Captures fat tails and tail dependence in market returns
    - Accumulated over 10-day horizon to capture path dependency

    3. PORTFOLIO REVALUATION
    - Full Revaluation: Closed-form pricing for equities, bonds (duration/convexity), options (Black-Scholes)
    - Delta-Gamma: Taylor expansion approximation for sensitivity-based P&L
    
    4. RISK METRICS
    - VaR (Value-at-Risk): Quantile of loss distribution at 99.9% confidence
    - ES (Expected Shortfall): Mean of losses exceeding VaR threshold (coherent risk measure)
    - Time scaling: Square-root rule for 10D → 1Y extrapolation (√(252/10))

    5. CAPITAL ALLOCATION
    - Euler Principle: Marginal contribution = Mean P&L in tail scenarios
    - Ensures portfolio ES = Σ(Component ES) with diversification

    6. STRESS TESTING
    - Predefined shock scenarios (e.g., -40% equity, +200bps rates)
    - Shift mean of factor distributions and re-simulate
    
    7. COVAR ANALYSIS
    - Conditional VaR: Portfolio VaR | Position at its own VaR level
    - ΔCoVaR: Incremental systemic contribution (Adrian & Brunnermeier, 2016)
    """

        ws.cell(method_row + 1, 1, methodology_text)
        ws.cell(method_row + 1, 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(method_row + 1, 1).font = Font(size=9, name="Courier New")
        ws.merge_cells(f"A{method_row + 1}:F{method_row + 25}")

        # References
        ref_row = method_row + 27
        ws.cell(ref_row, 1, "Key References").font = Font(
            size=14, bold=True, color=self.colors["header"]
        )

        references = [
            "• Basel Committee on Banking Supervision (2019). Minimum capital requirements for market risk (FRTB)",
            "• J.P. Morgan (1996). RiskMetrics Technical Document (4th ed.)",
            "• Adrian, T., & Brunnermeier, M. K. (2016). CoVaR. American Economic Review, 106(7), 1705-1741.",
            "• McNeil, A. J., Frey, R., & Embrechts, P. (2015). Quantitative Risk Management (2nd ed.). Princeton University Press.",
        ]

        for i, ref in enumerate(references, ref_row + 1):
            ws.cell(i, 1, ref)
            ws.cell(i, 1).font = Font(size=9, italic=True)
            ws.merge_cells(f"A{i}:F{i}")

        # Column widths
        ws.column_dimensions["A"].width = 80
        autofit_columns(ws)


def generate_market_risk_report(
    config: Dict[str, Any],
    engine: MarketRiskEconomicCapital,
    results: Dict[str, Any],
    output_dir: str = "reports",
) -> Path:
    return MarketRiskReporter(
        engine, results, config, output_dir
    ).generate_full_report()
