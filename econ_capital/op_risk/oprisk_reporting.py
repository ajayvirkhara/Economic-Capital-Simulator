"""
Regulatory-grade reporting for OpRisk Economic Capital & Stress Testing.
"""

from __future__ import annotations
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from .stress_tests import OpRiskStressTester, StressTestResult

from econ_capital.reporting_utils import apply_clean_style, autofit_columns


class OpRiskReporter:
    def __init__(
        self,
        tester: OpRiskStressTester,
        results: List[StressTestResult],
        config: Dict[str, Any],
        output_dir: str | Path = "reports",
    ):
        self.tester = tester
        self.results = sorted(results, key=lambda x: x.uplift_factor, reverse=True)
        self.config = config
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.filename = (
            self.output_dir
            / f"OpRisk_Stress_Test_Report_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        )

        self.colors = self.config.get("reporting_style", {}).get(
            "colors", self._default_colors()
        )
        # Ensure fallback for required keys if a partial config was loaded
        default_colors = self._default_colors()
        if "gold" not in self.colors:
            self.colors["gold"] = default_colors["gold"]
        if "header" not in self.colors:
            self.colors["header"] = default_colors["header"]

    def _default_colors(self):
        """Standard hardcoded colors for reporting style used as a fallback."""
        # Colors must be specified as hex codes (RRGGBB)
        return {
            "header": "1F4E78",  # Dark Blue/Grey for headers (Used for text and fill)
            "table_header": "DDEBF7",  # Light Blue/Grey fill
            "gold": "FFD700",  # Gold for top ranks (Used in _create_top10_sheet)
            "white": "FFFFFF",
        }

    def generate_full_report(self) -> Path:
        wb = Workbook()
        wb.remove(wb.active)

        self._create_cover_sheet(wb)
        self._create_summary_sheet(wb)
        self._create_results_sheet(wb)
        self._create_waterfall_sheet(wb)
        self._create_expert_scenarios_sheet(wb)

        wb.save(self.filename)
        print(f"Report generated: {self.filename}")
        return self.filename

    def _create_cover_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Cover", 0)
        lines = [
            "OPERATIONAL RISK",
            "ECONOMIC CAPITAL",
            "STRESS TEST REPORT",
            "",
            f"Run Date: {datetime.now():%Y-%m-%d %H:%M}",
            f"Baseline Capital: £{self.tester.baseline_capital:,.0f}",
            f"Scenarios Tested: {len(self.results)}",
        ]
        for i, line in enumerate(lines, 2):
            cell = ws[f"A{i}"]
            cell.value = line
            cell.font = Font(
                size=24 if i <= 4 else 14, bold=True, color=self.colors["header"]
            )
            cell.alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:F4")
        autofit_columns(ws)

    def _create_summary_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Summary", 1)
        ws["A1"] = "Executive Summary"
        ws["A1"].font = Font(size=18, bold=True, color=self.colors["header"])

        data = [
            ("Run Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("Baseline Capital (99.9% VaR)", f"£{self.tester.baseline_capital:,.0f}"),
            ("Number of Scenarios", len(self.results)),
            ("Worst-Case Uplift", f"{self.results[0].uplift_factor:.2f}x"),
            ("Worst-Case Scenario", self.results[0].name),
        ]
        for i, (k, v) in enumerate(data, 3):
            ws[f"A{i}"] = k
            ws[f"B{i}"] = v
            ws[f"A{i}"].font = Font(bold=True)
        autofit_columns(ws)

    def _create_results_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Detailed Results", 2)
        headers = [
            "Rank",
            "Scenario",
            "Description",
            "Base",
            "Stressed",
            "Uplift £",
            "Uplift ×",
            "Uplift %",
            "Runtime s",
        ]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.fill = PatternFill("solid", self.colors["header"])
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        for i, r in enumerate(self.results, 1):
            row = [
                i,
                r.name,
                r.description or "—",
                f"£{r.capital_base:,.0f}",
                f"£{r.capital_stressed:,.0f}",
                f"£{r.absolute_uplift:,.0f}",
                f"{r.uplift_factor:.2f}x",
                f"{r.uplift_pct:.1%}",
                f"{r.runtime_sec:.3f}s",
            ]
            for c, val in enumerate(row, 1):
                ws.cell(i + 1, c, val)
                # Apply gold highlight to top 3 contributors
                if i <= 3:
                    ws.cell(i + 1, c).fill = PatternFill("solid", self.colors["gold"])

        if len(self.results) > 0:
            last_row = len(self.results) + 1
            last_col = get_column_letter(len(headers))
            table_range = f"A1:{last_col}{last_row}"
            tab = Table(displayName="Results", ref=table_range)
            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9", showRowStripes=True
            )
            ws.add_table(tab)
        else:
            ws["A2"] = "No stress test results available"
        autofit_columns(ws)

    def _create_waterfall_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Waterfall", 3)
        ws["A1"] = "Capital Uplift Waterfall"
        ws["A1"].font = Font(size=16, bold=True, color=self.colors["header"])

        base = self.tester.baseline_capital
        top10 = self.results[:10]
        data = [["Baseline", base]] + [[r.name[:30], r.absolute_uplift] for r in top10]

        for r, (name, val) in enumerate(data, 2):
            ws[f"A{r}"] = name
            ws[f"B{r}"] = float(val)
            ws[f"B{r}"].number_format = "£#,##0"

        chart = BarChart()
        chart.title = "Top 10 Capital Impacts"
        chart.height = 14
        chart.width = 24
        data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(data))
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(data))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "D2")

        apply_clean_style(chart, "Capital (£)", num_points=len(data))
        autofit_columns(ws)

    def _create_expert_scenarios_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Expert Scenarios", 5)
        ws["A1"] = "Expert Judgment Scenario Capital"
        ws["A1"].font = Font(size=16, bold=True, color="1F4E78")

        details = self.config.get("expert_scenario_details", [])
        row = 3
        if details:
            headers = ["Scenario", "Probability", "Impact (£)", "Annual EL (£)"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row, col, h)
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", "DDEBF7")

            row += 1
            for d in details:
                ws.cell(row, 1, d["name"])
                ws.cell(row, 2, f"{d['probability']:.1%}")
                ws.cell(row, 3, f"£{d['impact']:,.0f}").number_format = "£#,##0"
                ws.cell(row, 4, f"£{d['annual_el']:,.0f}").number_format = "£#,##0"
                row += 1

            ws.cell(row, 3, "TOTAL EL").font = Font(bold=True)
            ws.cell(
                row, 4, f"£{self.config.get('expert_scenario_el', 0):,.0f}"
            ).number_format = "£#,##0"
            row += 1
            ws.cell(row, 3, "SCENARIO CAPITAL (×20)").font = Font(
                bold=True, color="C00000"
            )
            ws.cell(
                row, 4, f"£{self.config.get('expert_scenario_capital', 0):,.0f}"
            ).number_format = "£#,##0"
        else:
            ws["A3"] = "No expert judgment scenarios defined"

        for col in "ABCD":
            ws.column_dimensions[col].width = 25
        autofit_columns(ws)


def generate_oprisk_report(
    tester: OpRiskStressTester,
    results: List[StressTestResult],
    config: Dict[str, Any],
    output_dir: str | Path = "reports",
) -> Path:
    """
    Generate the Excel report using already-computed tester and stress test results.
    """
    return OpRiskReporter(
        tester=tester,
        results=results,
        config=config,
        output_dir=output_dir,
    ).generate_full_report()
