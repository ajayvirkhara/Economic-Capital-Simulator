"""
Regulatory-grade reporting for Credit Risk Economic Capital.
"""

from __future__ import annotations
from datetime import datetime
from pathlib import Path
from typing import Dict, Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .exposure_engine import ExposureEngine


class CreditRiskReporter:
    def __init__(
        self,
        engine: ExposureEngine,
        results: Dict[str, Any],
        config: Dict[str, Any],
        output_dir: str | Path | None = None,
    ):
        self.engine = engine
        self.results = results
        self.config = config
        if output_dir is None:
            # Resolves to econ_capital/credit_risk/reports
            pkg_dir = Path(__file__).resolve().parent
            self.output_dir = pkg_dir / "reports"
        else:
            self.output_dir = Path(output_dir)

        self.output_dir.mkdir(parents=True, exist_ok=True)

        self.filename = (
            self.output_dir
            / f"CreditRisk_EC_Report_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        )
        self.colors = self.config.get("reporting_style", {}).get(
            "colors", self._default_colors()
        )
        # Ensure fallback for required keys
        default_colors = self._default_colors()
        if "gold" not in self.colors:
            self.colors["gold"] = default_colors["gold"]
        if "header" not in self.colors:
            self.colors["header"] = default_colors["header"]

    def _default_colors(self):
        """Standard hardcoded colors for reporting style used as a fallback."""
        return {
            "header": "1F4E78",  # Dark Blue/Grey for headers
            "table_header": "DDEBF7",  # Light Blue/Grey fill
            "gold": "FFD700",  # Gold for top ranks
            "white": "FFFFFF",
        }

    def generate_full_report(self) -> Path:
        wb = Workbook()
        wb.remove(wb.active)

        self._create_cover_sheet(wb)
        self._create_summary_sheet(wb)
        self._create_waterfall_sheet(wb)
        self._create_full_portfolio_sheet(wb)

        wb.save(self.filename)
        print(f"Report generated: {self.filename}")
        return self.filename

    @property
    def baseline_capital(self) -> float:
        return float(self.results.get("baseline_capital", 0))

    def _create_cover_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Cover", 0)
        lines = [
            "CREDIT RISK",
            "ECONOMIC CAPITAL",
            "REPORT",
            "",
            f"Run Date: {datetime.now():%Y-%m-%d %H:%M}",
            f"Expected Loss (EL): £{self.results['EL_total']:,.0f}",
            f"Economic Capital (EC): £{self.results['EC_total']:,.0f}",
        ]
        for i, line in enumerate(lines, 2):
            cell = ws[f"A{i}"]
            cell.value = line
            cell.font = Font(
                size=24 if i <= 4 else 14, bold=True, color=self.colors["header"]
            )
            cell.alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:F4")

    def _create_summary_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Summary", 1)
        ws["A1"] = "Executive Summary"
        ws["A1"].font = Font(size=18, bold=True, color=self.colors["header"])

        data = [
            ("Run Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("Expected Loss (EL)", f"£{self.results['EL_total']:,.0f}"),
            ("Unexpected Loss (UL)", f"£{self.results['UL_total']:,.0f}"),
            ("Economic Capital (EC)", f"£{self.results['EC_total']:,.0f}"),
            ("Number of Paths", self.config.get("n_paths", "N/A")),
            (
                "Confidence Level",
                f"{self.config.get('confidence_level', 0.999) * 100:.1f}%",
            ),
        ]
        for i, (k, v) in enumerate(data, 3):
            ws[f"A{i}"] = k
            ws[f"B{i}"] = v
            ws[f"A{i}"].font = Font(bold=True)

    def _create_results_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Detailed Results", 2)
        headers = [
            "Metric",
            "Value",
        ]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.fill = PatternFill("solid", self.colors["header"])
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        data = [
            ("Expected Loss (EL)", f"£{self.results['EL_total']:,.0f}"),
            ("Unexpected Loss (UL)", f"£{self.results['UL_total']:,.0f}"),
            ("Economic Capital (EC)", f"£{self.results['EC_total']:,.0f}"),
        ]
        for i, (k, v) in enumerate(data, 2):
            ws.cell(i, 1, k)
            ws.cell(i, 2, v)

        if len(data) > 0:
            last_row = len(data) + 1
            last_col = get_column_letter(len(headers))
            table_range = f"A1:{last_col}{last_row}"
            tab = Table(displayName="DetailedResults", ref=table_range)
            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9", showRowStripes=True
            )
            ws.add_table(tab)
        else:
            ws["A2"] = "No results available"

    def _create_waterfall_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Waterfall", 3)
        ws["A1"] = "Capital Uplift Waterfall"
        ws["A1"].font = Font(size=16, bold=True, color=self.colors["header"])

        baseline = self.results.get("baseline_capital", 0)
        breakdown = self.results["capital_breakdown"]
        top10_positions = breakdown.head(10).index.tolist()

        data = [["Baseline", baseline]]
        for pos in top10_positions:
            uplift = breakdown.get(pos, 0)
            data.append([pos, uplift])

        for r, (name, val) in enumerate(data, 2):
            ws[f"A{r}"] = name
            ws[f"B{r}"] = val
            ws[f"B{r}"].number_format = "£#,##0"

        chart = BarChart()
        chart.title = "Top 10 Capital Impacts"
        chart.y_axis.title = "Capital (£)"
        chart.height = 14
        chart.width = 24
        data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(data))
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(data))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "D2")

    def _create_full_portfolio_sheet(self, wb: Workbook):
        """Creates a detailed table of all counterparties."""
        ws = wb.create_sheet("Full Portfolio", 3)
        ws["A1"] = "Detailed Counterparty Statistics"
        ws["A1"].font = Font(size=16, bold=True, color=self.colors["header"])

        if "full_data" not in self.results:
            ws["A3"] = "Detailed data not provided in results."
            return

        df = self.results["full_data"]

        # Define columns to display
        cols = ["name", "Sector", "EAD", "PD", "LGD", "EL", "UL", "EC_Marginal"]
        # Filter strictly for cols that exist in the dataframe
        cols = [c for c in cols if c in df.columns]

        # Write Headers
        for c_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=3, column=c_idx, value=col_name)
            cell.fill = PatternFill("solid", self.colors["header"])
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Write Data Rows
        for r_idx, row_data in enumerate(df[cols].itertuples(index=False), 4):
            for c_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                # Formatting based on column type
                header = cols[c_idx - 1]
                if isinstance(val, (int, float)):
                    if "PD" == header:
                        cell.number_format = "0.00%"
                    elif "LGD" == header:
                        cell.number_format = "0.00"
                    elif header in ["EAD", "EL", "UL", "EC_Marginal"]:
                        cell.number_format = "#,##0"

        # Add Excel Table feature
        last_col = get_column_letter(len(cols))
        last_row = len(df) + 3
        tab = Table(displayName="FullPortfolio", ref=f"A3:{last_col}{last_row}")
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
        )
        ws.add_table(tab)


def generate_creditrisk_report(
    config: Dict[str, Any],
    engine: ExposureEngine,
    results: Dict[str, Any],
    output_dir: str = "reports",
) -> Path:
    return CreditRiskReporter(
        engine, results, config, output_dir
    ).generate_full_report()
