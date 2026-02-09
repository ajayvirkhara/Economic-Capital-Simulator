"""
Firm-Wide Economic Capital Report — Regulatory-grade Excel output
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Dict, Any
from scipy.stats import t

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from econ_capital.reporting_utils import apply_clean_style, autofit_columns
from econ_capital.aggregate import aggregate_economic_capital


class FirmWideECReporter:
    def __init__(
        self,
        aggregated_results: Dict[str, Any],
        copula_df: float = 3.0,
        output_dir: str | Path = "reports",
    ):
        self.results = aggregated_results
        self.market_details = self.results.get("market_details", {})
        self.credit_details = self.results.get("credit_details", {})
        self.op_details = self.results.get("op_details", {})
        self.correlations = self.results.get("correlations", {})
        self.output_dir = Path(output_dir) / "firmwide"
        self.output_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.timestamp = timestamp  # Added for unique table names
        self.filename = self.output_dir / f"FirmWide_EC_Report_{timestamp}.xlsx"

        # Color scheme
        self.colors = {
            "header": "1F4E78",
            "table_header": "DDEBF7",
            "gold": "FFD700",
            "light_gold": "FFF2CC",
            "white": "FFFFFF",
        }
        self.copula_df = (
            copula_df if copula_df is not None else aggregated_results.get("copula_df")
        )

        if self.copula_df is not None:
            try:
                self.copula_df = float(self.copula_df)
            except (TypeError, ValueError):
                print(
                    f"Warning: Could not convert copula_df to float: {self.copula_df!r}"
                )
                self.copula_df = None

        self.correlation_matrix = self.results.get("correlation_matrix_array")
        self.correlation_regime = self.results.get("correlation_regime", "Static")

        # Convert corr matrix to numpy array if stored as list
        if self.correlation_matrix is not None and not isinstance(
            self.correlation_matrix, np.ndarray
        ):
            self.correlation_matrix = np.array(self.correlation_matrix)

    def generate_report(self) -> Path:
        wb = Workbook()
        wb.remove(wb.active)

        self._create_cover_sheet(wb)
        self._create_summary_sheet(wb)
        self._create_risk_contributions_sheet(wb)
        self._create_marginal_waterfall_sheet(wb)
        self._create_detailed_market_sheet(wb)
        self._create_market_risk_deep_dive_sheet(wb)
        self._create_detailed_credit_sheet(wb)
        self._create_detailed_oprisk_sheet(wb)
        self._create_correlation_matrix_sheet(wb)
        self._create_sensitivity_analysis_sheet(wb)

        wb.save(self.filename)
        wb.close()  # Added to prevent potential file corruption
        print("\nDetailed firm-wide Economic Capital report generated:")
        print(f"   {self.filename}")
        return self.filename

    def _create_cover_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Cover", 0)

        # Protect sheet for regulatory compliance
        ws.protection.sheet = True
        ws.protection.password = "ec_report"

        lines = [
            "FIRM-WIDE",
            "ECONOMIC CAPITAL",
            "REPORT",
            f"t-Copula (ν = {float(self.copula_df):.1f}) – 99.9% Confidence Level",
            "",
            f"Run Date: {datetime.now():%Y-%m-%d %H:%M}",
            f"Total Economic Capital (99.9%): £{self.results['EC_total']:,.0f}",
            f"Diversification Benefit: £{self.results['diversification_benefit']:,.0f}",
        ]

        for i, line in enumerate(lines, 2):
            cell = ws[f"A{i}"]
            cell.value = line
            cell.font = Font(
                size=28 if i <= 3 else 16,
                bold=True,
                color=self.colors["header"],
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("A2:G6")
        autofit_columns(ws)

    def _create_summary_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Executive Summary", 1)

        # Protect sheet for regulatory compliance
        ws.protection.sheet = True
        ws.protection.password = "ec_report"

        ws["A1"] = "Firm-Wide Economic Capital Summary"
        ws["A1"].font = Font(size=18, bold=True, color=self.colors["header"])

        copula_df = self.copula_df if self.copula_df is not None else 3.0
        t_quantile = t.ppf(0.999, copula_df)
        standalone_per_risk = {}
        for risk, vals in self.results["individual_risks"].items():
            if "Total_Standalone" in vals:
                standalone_per_risk[risk] = float(vals["Total_Standalone"])
            else:
                el = float(vals.get("EL", 0))
                ul = float(vals.get("UL", 0))
                standalone_per_risk[risk] = el + t_quantile * ul

        data = [
            ("Run Timestamp", self.results["run_timestamp"]),
            ("Total Expected Loss (EL)", f"£{self.results['EL_total']:,.0f}"),
            ("Portfolio Unexpected Loss (UL)", f"£{self.results['UL_portfolio']:,.0f}"),
            ("Total Economic Capital (99.9%)", f"£{self.results['EC_total']:,.0f}"),
            (
                "Diversification Benefit",
                f"£{self.results['diversification_benefit']:,.0f}",
            ),
            (
                "Standalone Sum (No Diversification)",
                f"£{self.results['EC_total'] + self.results['diversification_benefit']:,.0f}",
            ),
            ("Standalone Market EC", f"£{standalone_per_risk.get('Market', 0):,.0f}"),
            ("Standalone Credit EC", f"£{standalone_per_risk.get('Credit', 0):,.0f}"),
            ("Standalone OpRisk EC", f"£{standalone_per_risk.get('OpRisk', 0):,.0f}"),
        ]

        for i, (label, value) in enumerate(data, 3):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = value
            ws[f"A{i}"].font = Font(bold=True)
            if label != "Run Timestamp":
                ws[f"B{i}"].number_format = '"£"#,##0'
        ws.append(
            [
                "Note",
                "Marginal contributions are post-diversification and do not sum exactly to total EC",
            ]
        )
        autofit_columns(ws)

    def _create_risk_contributions_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Risk Contributions", 2)

        # Protect sheet for regulatory compliance
        ws.protection.sheet = True
        ws.protection.password = "ec_report"

        ws["A1"] = "Individual Risk Contributions"
        ws["A1"].font = Font(size=16, bold=True, color=self.colors["header"])

        # Use the pre-computed standalone values from aggregation (already scaled correctly)
        ind_risks = self.results.get("individual_risks", {})
        if len(ind_risks) != 3:
            ws["A3"] = (
                f"Warning: Only {len(ind_risks)} risk types found — expected 3 (Market, Credit, OpRisk)"
            )
        else:
            df = pd.DataFrame(
                {
                    "Expected Loss (EL)": [ind_risks[r]["EL"] for r in ind_risks],
                    "Unexpected Loss (UL)": [ind_risks[r]["UL"] for r in ind_risks],
                    "Standalone EC": [
                        ind_risks[r]["Total_Standalone"] for r in ind_risks
                    ],
                },
                index=list(ind_risks.keys()),
            )

            # Format as currency strings
            df = df.map(lambda x: f"£{float(x):,.0f}" if pd.notnull(x) else "£0")

        headers = [
            "Risk Type",
            "Expected Loss (EL)",
            "Unexpected Loss (UL)",
            "Standalone EC",
        ]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(3, c, h)
            cell.fill = PatternFill("solid", self.colors["header"])
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        for r, (risk, row) in enumerate(df.iterrows(), 4):
            ws.cell(r, 1, risk)
            for c, val in enumerate(row, 2):
                cell = ws.cell(r, c, val)
                cell.number_format = '"£"#,##0'

        # Format as table
        last_row = len(df) + 3
        last_col = get_column_letter(len(headers))
        tab = Table(
            displayName=f"RiskContributions_{self.timestamp}",
            ref=f"A3:{last_col}{last_row}",
        )
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showRowStripes=True
        )
        ws.add_table(tab)

        # --- ADD FOOTNOTE ---
        last_data_row = len(df) + 3  # 3 = header row + 2 for title/subtitle

        # Add footnote 2–3 rows below the table
        footnote_row = last_data_row + 3

        ws.cell(row=footnote_row, column=1).value = "Footnote:"
        ws.cell(row=footnote_row, column=1).font = Font(
            bold=True, size=12, color="1F4E78"
        )

        footnote_text = (
            "Individual ULs and Standalone EC are calculated on a standalone basis using the t-copula "
            "(ν=3) 99.9% quantile (≈10.215). Portfolio UL is the std dev of simulated aggregate losses "
            "(reflects correlations and diversification). They are not expected to sum directly."
        )

        ws.cell(row=footnote_row, column=2).value = footnote_text
        ws.cell(row=footnote_row, column=2).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        ws.cell(row=footnote_row, column=2).font = Font(
            size=10, italic=True, color="555555"
        )

        # Merge footnote across columns
        ws.merge_cells(
            start_row=footnote_row,
            start_column=2,
            end_row=footnote_row + 2,
            end_column=6,
        )

        autofit_columns(ws)

    def _create_marginal_waterfall_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Marginal Contributions", 3)

        # Protect sheet for regulatory compliance
        ws.protection.sheet = True
        ws.protection.password = "ec_report"

        ws["A1"] = "Marginal Economic Capital Contributions"
        ws["A1"].font = Font(size=16, bold=True, color=self.colors["header"])

        marginal = self.results["marginal_contributions"]
        df = (
            pd.Series(marginal).sort_values(ascending=False).to_frame("Marginal EC (£)")
        )

        # Write data
        ws["A3"] = "Risk Type"
        ws["B3"] = "Marginal Contribution"
        for cell in ws[3]:
            cell.fill = PatternFill("solid", self.colors["header"])
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        for r, (risk, contrib) in enumerate(df.itertuples(), 4):
            ws.cell(r, 1, risk)
            ws.cell(r, 2, contrib)
            ws.cell(r, 2).number_format = '"£"#,##0'
            if r <= 6:  # Top 3 gold highlight
                for c in [1, 2]:
                    ws.cell(r, c).fill = PatternFill(
                        "solid",
                        self.colors["gold"] if r <= 5 else self.colors["light_gold"],
                    )

        # Bar chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Marginal Economic Capital by Risk Type"
        chart.x_axis.title = "Risk Type"

        # Data
        data = Reference(ws, min_col=2, min_row=4, max_row=3 + len(df))
        cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(df))
        chart.add_data(data, titles_from_data=False)
        if chart.series:
            chart.series[0].name = ""  # Explicitly set series name to empty string
        chart.set_categories(cats)

        apply_clean_style(chart, "Marginal EC (£)")
        ws.add_chart(chart, "E2")

        # Table
        last_row = len(df) + 3
        tab = Table(
            displayName=f"MarginalContributions_{self.timestamp}", ref=f"A3:B{last_row}"
        )
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True
        )
        ws.add_table(tab)

        # --- ADD FOOTNOTE ---
        footnote_row = last_row + 3

        ws.cell(footnote_row, 1).value = "Footnote:"
        ws.cell(footnote_row, 1).font = Font(bold=True, color="1F4E78")

        footnote_text = (
            "Marginal Economic Capital contributions represent each risk type's incremental contribution "
            "to the firm-wide diversified Economic Capital at 99.9% confidence level (t-copula, ν=3). "
            "By construction (Euler allocation principle), the sum of marginal contributions is (approximately) "
            "equal to the total diversified Economic Capital (£{:,}). Minor differences may occur due to "
            "Monte-Carlo simulation noise."
        ).format(int(round(self.results["EC_total"])))

        ws.cell(footnote_row, 2).value = footnote_text
        ws.cell(footnote_row, 2).alignment = Alignment(wrapText=True, vertical="top")
        ws.cell(footnote_row, 2).font = Font(size=10, italic=True, color="444444")

        # Merge across reasonable width
        ws.merge_cells(
            start_row=footnote_row,
            start_column=2,
            end_row=footnote_row + 2,
            end_column=6,
        )

        autofit_columns(ws)

    def _create_detailed_market_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Detailed Market Risk", 4)

        # Protect sheet for regulatory compliance
        ws.protection.sheet = True
        ws.protection.password = "ec_report"

        ws["A1"] = "Detailed Market Risk Breakdown"
        ws["A1"].font = Font(size=16, bold=True, color=self.colors["header"])

        standalone_market = (
            self.results.get("individual_risks", {})
            .get("Market", {})
            .get("Total_Standalone", 0)
        )

        if standalone_market == 0:
            # Fallback to es_1y_999 if aggregation didn't populate
            standalone_market = (
                self.market_details.get("es_1y_999", 0) * 3.3
            )  # approx t(3) uplift

        ws["A2"] = (
            f"Standalone Market Risk EC (Pre-Divesification): £{int(round(standalone_market)):,.0f}"
        )
        ws["A2"].font = Font(bold=True, size=12, color="1F4E78")

        # Top 10 contributors
        breakdown = self.market_details.get("capital_breakdown", pd.Series())
        if breakdown.empty:
            ws["A3"] = "No market risk details available."
            return

        raw_sum = (
            breakdown.values.sum()
            if hasattr(breakdown, "values")
            else sum(breakdown.values())
        )
        scale_factor = 1.0
        if raw_sum > 0 and abs(raw_sum - standalone_market) > 1000:
            scale_factor = standalone_market / raw_sum

        # Apply scaling safely
        if isinstance(breakdown, pd.DataFrame):
            breakdown_scaled = breakdown.iloc[:, 0] * scale_factor
        else:
            breakdown_scaled = breakdown * scale_factor

        headers = ["Rank", "Position", "Capital Contribution (£)"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(3, c, h)
            cell.fill = PatternFill("solid", self.colors["header"])
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        top_n = breakdown_scaled.head(10)
        # Define variables for row tracking
        start_top10 = 4

        for i, (pos, contrib) in enumerate(top_n.items(), start_top10):
            ws.cell(i, 1, i - 3)
            ws.cell(i, 2, pos)
            ws.cell(i, 3, float(contrib)).number_format = '"£"#,##0'
            if i - 3 <= 3:
                for c in range(1, 4):
                    ws.cell(i, c).fill = PatternFill("solid", self.colors["gold"])

            if i - 3 <= 3:  # Highlight top 3
                for c in range(1, 4):
                    ws.cell(i, c).fill = PatternFill("solid", self.colors["gold"])

        num_items = len(top_n)
        if num_items > 0:
            last_row = start_top10 + num_items - 1
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Top Market Capital Impacts (Euler)"
            chart.x_axis.title = "Position"

            # Data and categories
            data = Reference(ws, min_col=3, min_row=start_top10, max_row=last_row)
            cats = Reference(ws, min_col=2, min_row=start_top10, max_row=last_row)
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(cats)

            # Y-axis formatting
            chart.y_axis.number_format = '"£"#,##0'

            apply_clean_style(chart, "Capital Contribution (£)")
            ws.add_chart(chart, "E2")

            # Table
            tab = Table(
                displayName=f"MarketTop10_{self.timestamp}", ref=f"A3:C{last_row}"
            )
            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showRowStripes=True
            )
            ws.add_table(tab)

        note_row = ws.max_row + 2  # 2 rows below the last data row
        ws.cell(row=note_row, column=1).value = (
            "Note: Sum of detailed market positions is standalone (pre-diversification). "
            "Marginal contribution sheet reflects post-diversification effect in the firm-wide portfolio."
        )
        ws.cell(row=note_row, column=1).font = Font(italic=True, color="555555")
        ws.merge_cells(
            start_row=note_row,
            start_column=1,
            end_row=note_row,
            end_column=ws.max_column,
        )
        autofit_columns(ws)

    def _create_detailed_credit_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Detailed Credit Risk", 5)

        # Protect sheet for regulatory compliance
        ws.protection.sheet = True
        ws.protection.password = "ec_report"

        ws["A1"] = "Detailed Credit Risk Portfolio"
        ws["A1"].font = Font(size=16, bold=True, color=self.colors["header"])

        standalone_credit = (
            self.results.get("individual_risks", {})
            .get("Credit", {})
            .get("Total_Standalone", 0)
        )

        if standalone_credit == 0:
            # Fallback: reconstruct from normalized UL + t(3) quantile
            credit_norm = self.results.get("individual_risks", {}).get("Credit", {})
            if credit_norm:
                t_quant = t.ppf((0.999, 3), self.copula_df or 3.0)
                standalone_credit = credit_norm.get(
                    "EL", 0
                ) + t_quant * credit_norm.get("UL", 0)

        ws["A2"] = (
            f"Standalone Credit Risk EC (Pre-Diversification): £{int(round(standalone_credit)):,.0f}"
        )
        ws["A2"].font = Font(bold=True, size=12, color="1F4E78")

        # Full portfolio table
        df = self.credit_details.get("full_data", pd.DataFrame())
        if df.empty:
            ws["A3"] = "No credit risk details available."
            return

        if "EC_Marginal" in df.columns:
            current_sum = df["EC_Marginal"].sum()
            if current_sum > 0 and abs(current_sum - standalone_credit) > 1000:
                scale_factor = standalone_credit / current_sum
                # Work on a copy to avoid side effects
                df = df.copy()
                df["EC_Marginal"] = df["EC_Marginal"] * scale_factor
            else:
                scale_factor = 1.0
        else:
            scale_factor = 1.0

        cols_map = {
            "name": "Counterparty",
            "Sector": "Sector",
            "EAD": "EAD",
            "PD": "PD",
            "LGD": "LGD",
            "EL": "EL",
            "UL": "UL",
            "EC_Marginal": "Credit PF Contrib",
        }
        available_cols = [c for c in cols_map.keys() if c in df.columns]

        # Write headers
        for c_idx, col_key in enumerate(available_cols, 1):
            # Use mapped name if available, else raw name
            cell = ws.cell(row=3, column=c_idx, value=cols_map[col_key])
            cell.fill = PatternFill("solid", self.colors["header"])
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Write data rows
        for r_idx, row in enumerate(df[available_cols].itertuples(index=False), 4):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                header = available_cols[c_idx - 1]
                if isinstance(val, (int, float)):
                    if header == "PD":
                        cell.number_format = "0.00%"
                    elif header == "LGD":
                        cell.number_format = "0.00"
                    elif header in ["EAD", "EL", "UL", "EC_Marginal"]:
                        cell.number_format = '"£"#,##0'

        # Table
        last_row = len(df) + 3
        last_col = get_column_letter(len(available_cols))
        tab = Table(
            displayName=f"CreditPortfolio_{self.timestamp}",
            ref=f"A3:{last_col}{last_row}",
        )
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True
        )
        ws.add_table(tab)

        credit_note_row = last_row + 3
        ws.cell(
            credit_note_row, 1
        ).value = "Note: 'Credit PF Contrib' is the marginal contribution to the standalone (pre-diversification) Credit Portfolio EC. "
        ws.cell(credit_note_row, 1).font = Font(italic=True, color="555555")
        ws.merge_cells(
            start_row=credit_note_row,
            start_column=1,
            end_row=credit_note_row + 3,
            end_column=8,
        )

        autofit_columns(ws)

    def _create_detailed_oprisk_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Detailed Op Risk", 6)

        # Protect sheet for regulatory compliance
        ws.protection.sheet = True
        ws.protection.password = "ec_report"

        ws["A1"] = "Detailed Op Risk Scenarios"
        ws["A1"].font = Font(size=16, bold=True, color=self.colors["header"])

        raw_results = (
            self.op_details.get("stress_test_results", [])
            if isinstance(self.op_details, dict)
            else []
        )
        expert_details = (
            self.op_details.get("expert_scenario_details", [])
            if isinstance(self.op_details, dict)
            else []
        )

        standalone_oprisk = (
            self.results.get("individual_risks", {})
            .get("OpRisk", {})
            .get("Total_Standalone", 0)
        )
        ws["A2"] = (
            f"Standalone OpRisk EC (99.9% t-copula; pre-diversification): £{int(round(standalone_oprisk)):,.0f}"
        )
        ws["A2"].font = Font(bold=True, size=12, color="1F4E78")

        if not raw_results:
            ws["A3"] = "No op risk details available."

        else:
            # Helper to safely get attributes whether objects or dicts
            def get_attr(obj, attr, default=0):
                if isinstance(obj, dict):
                    return obj.get(attr, default)
                return getattr(obj, attr, default)

            # Sort and take top 10
            top10 = sorted(
                raw_results, key=lambda x: get_attr(x, "uplift_factor"), reverse=True
            )[:10]

            headers = [
                "Rank",
                "ScenarioID",
                "Description",
                "Baseline Cap",
                "Stressed Capital",
                "Uplift",
            ]
            for c, h in enumerate(headers, 1):
                cell = ws.cell(3, c, h)
                cell.fill = PatternFill("solid", self.colors["header"])
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
                cell.number_format = '"£"#,##0'

            # Write data rows
            for i, r in enumerate(top10, 4):
                ws.cell(i, 1, int(i - 3))
                ws.cell(i, 2, str(get_attr(r, "name", "N/A")))
                ws.cell(i, 3, str(get_attr(r, "description", "---")))

                # Baseline capital
                base_cap = float(get_attr(r, "capital_base", 0.0))
                ws.cell(i, 4, base_cap).number_format = '"£"#,##0'

                # Stressed capital
                stressed_cap = float(get_attr(r, "capital_stressed", 0.0))
                ws.cell(i, 5, stressed_cap).number_format = '"£"#,##0'

                # Uplift as raw float
                uplift = float(get_attr(r, "uplift_factor", 0.0))
                ws.cell(i, 6, uplift).number_format = "0.00x"

                if i - 3 <= 3:  # Top 3 highlighting
                    for c in range(1, 7):
                        ws.cell(i, c).fill = PatternFill("solid", self.colors["gold"])

            # Table definition with strictly unique name and checked range
            last_row = len(top10) + 3
            if len(top10) > 0:
                tab = Table(
                    displayName=f"OpRiskTop10_{self.timestamp}", ref=f"A3:F{last_row}"
                )

                # Style
                style = TableStyleInfo(
                    name="TableStyleMedium2",
                    showRowStripes=True,
                )
                tab.tableStyleInfo = style

                # Add to worksheet
                ws.add_table(tab)

        # Expert scenarios
        if expert_details:
            start_expert = ws.max_row + 4
            ws.cell(start_expert, 1, "Expert Judgment Scenarios").font = Font(
                size=14, bold=True, color=self.colors["header"]
            )

            expert_headers = ["Scenario", "Probability", "Impact (£)", "Annual EL (£)"]
            for c, h in enumerate(expert_headers, 1):
                cell = ws.cell(start_expert + 1, c, h)
                cell.fill = PatternFill("solid", self.colors["header"])
                cell.font = Font(color="FFFFFF", bold=True)

            for i, d in enumerate(expert_details, start_expert + 2):
                ws.cell(i, 1, d.get("name", "N/A"))
                ws.cell(i, 2, f"{d.get('probability', 0.0):.1%}")
                ws.cell(i, 3, d.get("impact", 0.0)).number_format = '"£"#,##0'
                ws.cell(i, 4, d.get("annual_el", 0.0)).number_format = '"£"#,##0'

            # Add table for expert scenarios
            expert_last_row = start_expert + 1 + len(expert_details)
            expert_tab = Table(
                displayName=f"ExpertScenarios_{self.timestamp}",
                ref=f"A{start_expert + 1}:D{expert_last_row}",
            )
            expert_tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showRowStripes=True
            )
            ws.add_table(expert_tab)

        op_note_row = ws.max_row + 3
        ws.cell(op_note_row, 1).value = (
            "Note: 'Baseline Cap' represents the Standalone (pre-diversification) OpRisk Economic Capital. "
            "It is an unconditional 99.9% UL estimate derived from statistical modeling and does not include manual Expert Judgment overlays or firm-wide diversification credits."
        )
        ws.cell(op_note_row, 1).font = Font(italic=True, color="555555")
        ws.merge_cells(
            start_row=op_note_row, start_column=1, end_row=op_note_row + 2, end_column=6
        )

        autofit_columns(ws)

    def _create_correlation_matrix_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Correlation Matrix", 7)

        # Protect sheet for regulatory compliance
        ws.protection.sheet = True
        ws.protection.password = "ec_report"

        ws["A1"] = "Inter-Risk Correlation Matrix"
        ws["A1"].font = Font(size=16, bold=True, color=self.colors["header"])

        # Display regime information
        regime_info = self.correlations.get("regime", "Static")
        method_info = self.correlations.get("method", "Static")

        ws["A2"] = f"Regime: {regime_info} | Method: {method_info}"
        ws["A2"].font = Font(size=12, italic=True, color="1F4E78")
        ws.merge_cells("A2:D2")

        # Extract matrix (handle both old dict format and new matrix format)
        if "matrix" in self.correlations:
            # New format: numpy array stored as list
            corr_array = np.array(self.correlations["matrix"])
            risk_types = ["Market", "Credit", "OpRisk"]
        else:
            # Old format: nested dicts
            risk_types = list(self.correlations.keys())
            if not risk_types:
                ws["A3"] = "No correlation data available."
                return

            corr_array = np.array(
                [
                    [self.correlations.get(rt1, {}).get(rt2, 0.0) for rt2 in risk_types]
                    for rt1 in risk_types
                ]
            )

        # Correlation matrix
        ws.cell(4, 1, "Current Correlation Matrix").font = Font(
            size=14, bold=True, color=self.colors["header"]
        )

        # Headers
        ws.cell(5, 1, "Risk Type").font = Font(color="FFFFFF", bold=True)
        ws.cell(5, 1).fill = PatternFill("solid", self.colors["header"])

        for c, risk in enumerate(risk_types, 2):
            cell = ws.cell(5, c, risk)
            cell.fill = PatternFill("solid", self.colors["header"])
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Write data
        for r, rt1 in enumerate(risk_types, 6):
            ws.cell(r, 1, str(rt1)).font = Font(bold=True)
            for c, rt2 in enumerate(risk_types, 2):
                corr_val = float(corr_array[r - 6, c - 2])
                cell = ws.cell(r, c, corr_val)
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="center")

                # Conditional formatting: highlight high correlations
                if r != c:  # Off-diagonal
                    if abs(corr_val) > 0.5:
                        cell.fill = PatternFill(
                            "solid", fgColor="FFC7CE"
                        )  # Red for high
                    elif abs(corr_val) > 0.3:
                        cell.fill = PatternFill(
                            "solid", fgColor="FFEB9C"
                        )  # Yellow for moderate

        last_row = len(risk_types) + 5
        last_col = get_column_letter(len(risk_types) + 1)

        # Table for current matrix
        tab = Table(
            displayName=f"CorrMatrixTable_{self.timestamp}",
            ref=f"A5:{last_col}{last_row}",
        )

        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showRowStripes=True
        )

        ws.add_table(tab)

        # === STATIC BASELINE COMPARISON (if dynamic was used) ===
        if "static_fallback" in self.correlations:
            comparison_start = last_row + 3

            ws.cell(
                comparison_start, 1, "Static Baseline (For Comparison)"
            ).font = Font(size=14, bold=True, color="666666")

            static_corrs = self.correlations["static_fallback"]

            # Build static correlation array
            static_array = np.array(
                [
                    [
                        1.0,
                        static_corrs.get("Market", {}).get("Credit", 0.3),
                        static_corrs.get("Market", {}).get("OpRisk", 0.1),
                    ],
                    [
                        static_corrs.get("Credit", {}).get("Market", 0.3),
                        1.0,
                        static_corrs.get("Credit", {}).get("OpRisk", 0.2),
                    ],
                    [
                        static_corrs.get("OpRisk", {}).get("Market", 0.1),
                        static_corrs.get("OpRisk", {}).get("Credit", 0.2),
                        1.0,
                    ],
                ]
            )

            # Headers for static table
            ws.cell(comparison_start + 1, 1, "Risk Type").font = Font(
                bold=True, color="666666"
            )
            ws.cell(comparison_start + 1, 1).fill = PatternFill(
                "solid", fgColor="E0E0E0"
            )

            for c, risk in enumerate(risk_types, 2):
                cell = ws.cell(comparison_start + 1, c, risk)
                cell.fill = PatternFill("solid", fgColor="E0E0E0")
                cell.font = Font(bold=True, color="666666")
                cell.alignment = Alignment(horizontal="center")

            # Write static correlation data
            for r, rt1 in enumerate(risk_types, comparison_start + 2):
                ws.cell(r, 1, str(rt1)).font = Font(bold=True, color="666666")
                for c, rt2 in enumerate(risk_types, 2):
                    static_val = float(static_array[r - comparison_start - 2, c - 2])
                    cell = ws.cell(r, c, static_val)
                    cell.number_format = "0.00"
                    cell.alignment = Alignment(horizontal="center")
                    cell.font = Font(color="666666")

            static_last_row = len(risk_types) + comparison_start + 1

            # Table for static matrix
            static_tab = Table(
                displayName=f"CorrMatrixStatic_{self.timestamp}",
                ref=f"A{comparison_start + 1}:{last_col}{static_last_row}",
            )
            static_tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleLight11", showRowStripes=True
            )
            ws.add_table(static_tab)

            # === DELTA ANALYSIS (Current - Static) ===
            delta_start = static_last_row + 3

            ws.cell(
                delta_start, 1, "Correlation Change (Current - Static)"
            ).font = Font(size=14, bold=True, color=self.colors["header"])

            # Headers
            ws.cell(delta_start + 1, 1, "Risk Pair").font = Font(bold=True)
            ws.cell(delta_start + 1, 2, "Current").font = Font(bold=True)
            ws.cell(delta_start + 1, 3, "Static").font = Font(bold=True)
            ws.cell(delta_start + 1, 4, "Δ (Change)").font = Font(bold=True)
            ws.cell(delta_start + 1, 5, "Impact").font = Font(bold=True)

            for c in range(1, 6):
                ws.cell(delta_start + 1, c).fill = PatternFill(
                    "solid", self.colors["table_header"]
                )

            # Calculate and display deltas for off-diagonal pairs
            delta_row = delta_start + 2
            risk_pairs = [
                ("Market", "Credit", 0, 1),
                ("Market", "OpRisk", 0, 2),
                ("Credit", "OpRisk", 1, 2),
            ]

            for pair_name1, pair_name2, idx1, idx2 in risk_pairs:
                current_corr = float(corr_array[idx1, idx2])
                static_corr = float(static_array[idx1, idx2])
                delta = current_corr - static_corr

                ws.cell(delta_row, 1, f"{pair_name1}-{pair_name2}")
                ws.cell(delta_row, 2, current_corr).number_format = "0.00"
                ws.cell(delta_row, 3, static_corr).number_format = "0.00"

                delta_cell = ws.cell(delta_row, 4, delta)
                delta_cell.number_format = "+0.00;-0.00"

                # Color code delta
                if abs(delta) > 0.2:
                    delta_cell.fill = PatternFill(
                        "solid", fgColor="FFC7CE"
                    )  # Large change - red
                    impact = "High"
                elif abs(delta) > 0.1:
                    delta_cell.fill = PatternFill(
                        "solid", fgColor="FFEB9C"
                    )  # Moderate - yellow
                    impact = "Moderate"
                else:
                    delta_cell.fill = PatternFill(
                        "solid", fgColor="C6EFCE"
                    )  # Low - green
                    impact = "Low"

                ws.cell(delta_row, 5, impact)
                delta_row += 1

            # Add explanatory note
            note_row = delta_row + 2
            ws.cell(note_row, 1, "Note:").font = Font(
                bold=True, color=self.colors["header"]
            )
            ws.cell(
                note_row,
                2,
                "Positive Δ indicates increased correlation in current regime (higher contagion risk). "
                "High impact changes (|Δ| > 0.2) materially affect diversification benefits and "
                "firm-wide Economic Capital. Review correlation assumptions if changes exceed ±0.3.",
            )
            ws.cell(note_row, 2).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(note_row, 2).font = Font(size=9, italic=True, color="555555")
            ws.merge_cells(f"B{note_row}:E{note_row + 2}")

        autofit_columns(ws)

    def _create_sensitivity_analysis_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Sensitivity Analysis", 8)

        # Protect sheet for regulatory compliance
        ws.protection.sheet = True
        ws.protection.password = "ec_report"

        ws["A1"] = (
            f"Sensitivity to Confidence Level (t-copula ν={self.copula_df or 3.0:.1f})"
        )
        ws["A1"].font = Font(size=16, bold=True, color=self.colors["header"])

        levels = [0.99, 0.995, 0.999]

        # Write headers
        ws.cell(3, 1, "Confidence Level").font = Font(color="FFFFFF", bold=True)
        ws.cell(3, 2, "EC Total (£)").font = Font(color="FFFFFF", bold=True)
        for c in [1, 2]:
            ws.cell(3, c).fill = PatternFill("solid", self.colors["header"])

        # [Construct explicit OpRisk dictionary for aggregation
        baseline = self.op_details.get("baseline_metrics", {})
        op_results_for_agg = {
            "capital_999": baseline.get("capital_999", 0.0),
            "expected_loss": baseline.get("expected_loss", 0.0),
            # Pass total_capital if available at top level for scaling logic
            "total_capital": self.op_details.get(
                "total_capital", baseline.get("capital_999", 0.0)
            ),
        }

        # Write data rows
        for i, level in enumerate(levels, 4):
            # Pass full details (self.market_details, self.credit_details)
            _, _, EC_total, _, _ = aggregate_economic_capital(
                market_results=self.market_details,
                credit_results=self.credit_details,
                op_results=op_results_for_agg,
                confidence_level=level,
                copula_df=self.copula_df if self.copula_df is not None else 3.0,
                n_sim=750_000,
                correlation_matrix=self.correlation_matrix,
                correlation_regime=self.correlation_regime,
            )

            ws.cell(i, 1, f"{level * 100:.1f}%")
            cell_b = ws.cell(i, 2, EC_total)
            cell_b.number_format = '"£"#,##0'

        # Table
        last_row = len(levels) + 3
        tab = Table(displayName=f"Sensitivity_{self.timestamp}", ref=f"A3:B{last_row}")
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True
        )
        ws.add_table(tab)

        # Chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "EC Sensitivity to Confidence Level"
        chart.x_axis.title = "Confidence Level"
        chart.height = 12
        chart.width = 18

        # Data and Categories
        data = Reference(ws, min_col=2, min_row=4, max_row=last_row)
        cats = Reference(ws, min_col=1, min_row=4, max_row=last_row)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)

        # X-axis Labels
        chart.x_axis.title = "Confidence Level"

        apply_clean_style(chart, "Economic Capital (£)")

        ws.add_chart(chart, "E2")
        autofit_columns(ws)

    def _create_market_risk_deep_dive_sheet(self, wb: Workbook):
        """Comprehensive market risk metrics including CoVaR, full reval, stress tests."""
        ws = wb.create_sheet("Market Risk Deep Dive", 5)

        ws.protection.sheet = True
        ws.protection.password = "ec_report"

        ws["A1"] = "Market Risk - Advanced Metrics & Analysis"
        ws["A1"].font = Font(size=16, bold=True, color=self.colors["header"])

        # ========== SECTION 1: VaR/ES BREAKDOWN ==========
        ws["A3"] = "1. VaR/ES Metrics"
        ws["A3"].font = Font(size=14, bold=True, color=self.colors["header"])

        metrics_data = [
            ("10D VaR (99.9%)", self.market_details.get("var_10d_999", 0)),
            ("10D ES (99.9%)", self.market_details.get("es_10d_999", 0)),
            ("1Y VaR (99.9%)", self.market_details.get("var_1y_999", 0)),
            ("1Y ES (99.9%)", self.market_details.get("es_1y_999", 0)),
            ("Stressed VaR (1Y)", self.market_details.get("stressed_var_1y_999", 0)),
            ("Stressed ES (1Y)", self.market_details.get("stressed_es_1y_999", 0)),
        ]
        for i, (label, value) in enumerate(metrics_data, 4):
            ws.cell(i, 1, label).font = Font(bold=True)
            cell = ws.cell(i, 2, value)
            cell.number_format = "£#,##0"

            # Highlight stressed metrics
            if "Stressed" in label:
                cell.fill = PatternFill("solid", fgColor="FFC7CE")

        # ========== SECTION 2: COVAR METRICS ==========
        covar_start = 11
        ws.cell(covar_start, 1, "2. Systemic Risk (CoVaR)")
        ws.cell(covar_start, 1).font = Font(
            size=14, bold=True, color=self.colors["header"]
        )

        covar_metrics = self.market_details.get("covar_metrics", {})

        if covar_metrics:
            headers = ["Position", "ΔCoVaR (£)", "Systemic %"]
            for c, h in enumerate(headers, 1):
                cell = ws.cell(covar_start + 1, c, h)
                cell.fill = PatternFill("solid", self.colors["table_header"])
                cell.font = Font(bold=True)

            covar_df = pd.DataFrame(covar_metrics).T
            covar_df = covar_df.sort_values(
                "delta_covar", key=abs, ascending=False
            ).head(5)

            for i, (pos, row) in enumerate(covar_df.iterrows(), covar_start + 2):
                ws.cell(i, 1, pos)
                ws.cell(i, 2, float(row["delta_covar"])).number_format = "£#,##0"
                ws.cell(
                    i, 3, float(row["systemic_contribution_pct"]) / 100
                ).number_format = "0.00%"
        else:
            ws.cell(covar_start + 1, 1, "CoVaR analysis not available")
            ws.cell(covar_start + 1, 1).font = Font(italic=True, color="999999")

        # ========== SECTION 3: FULL REVALUATION INFO ==========
        reval_start = covar_start + 8
        ws.cell(reval_start, 1, "3. Valuation Method")
        ws.cell(reval_start, 1).font = Font(
            size=14, bold=True, color=self.colors["header"]
        )

        used_full_reval = self.market_details.get("used_full_revaluation", False)

        if used_full_reval:
            ws.cell(reval_start + 1, 1, "Pricing Method:")
            ws.cell(reval_start + 1, 2, "Full Revaluation (Closed-Form)")
            ws.cell(reval_start + 1, 2).font = Font(bold=True, color="008000")

            ws.cell(reval_start + 2, 1, "Coverage:")
            ws.cell(
                reval_start + 2,
                2,
                "Equities (spot), Bonds (duration/convexity), FX, Options (Black-Scholes)",
            )

            ws.cell(reval_start + 3, 1, "Benefits:")
            ws.cell(
                reval_start + 3,
                2,
                "Captures non-linear risk, path dependency, and convexity effects",
            )
        else:
            ws.cell(reval_start + 1, 1, "Pricing Method:")
            ws.cell(reval_start + 1, 2, "Delta-Gamma Approximation")
            ws.cell(reval_start + 1, 2).font = Font(bold=True, color="C00000")

            ws.cell(reval_start + 2, 1, "Note:")
            ws.cell(
                reval_start + 2,
                2,
                "Linear approximation may underestimate risk for options and convex instruments",
            )

        # ========== SECTION 4: MODEL DIAGNOSTICS ==========
        diag_start = reval_start + 6
        ws.cell(diag_start, 1, "4. Model Diagnostics").font = Font(
            size=14, bold=True, color=self.colors["header"]
        )

        has_hist_var = "historical_var_1y_999" in self.market_details

        if has_hist_var:
            hist_var = self.market_details["historical_var_1y_999"]
            param_var = self.market_details.get("var_1y_999", 0)
            if param_var > 0:
                diff_pct = (hist_var - param_var) / param_var * 100
                ws.cell(diag_start + 1, 1, "Historical vs Parametric 1Y VaR (99.9%):")
                ws.cell(
                    diag_start + 1, 2, f"{diff_pct:+.1f}%"
                ).number_format = "0.00%;[Red]-0.00%"

                if abs(diff_pct) < 10:
                    status = "Well calibrated ✓"
                    color = "008000"
                else:
                    status = "Review assumptions"
                    color = "C00000"
                ws.cell(diag_start + 2, 1, "Assessment:").font = Font(bold=True)
                ws.cell(diag_start + 2, 2, status).font = Font(color=color, bold=True)
        else:
            # Fallback content when historical is off
            ws.cell(diag_start + 1, 1, "Historical VaR comparison")
            ws.cell(diag_start + 1, 2, "Not enabled (use_historical_var = False)")
            ws.cell(diag_start + 1, 2).font = Font(italic=True, color="777777")

            ws.cell(
                diag_start + 3, 1, "→ Enable historical simulation for full diagnostics"
            )

        autofit_columns(ws)


def generate_firmwide_ec_report(
    aggregated_results: Dict[str, Any],
    copula_df: float = 3.0,
    output_dir: str | Path = "reports",
) -> Path:
    """Convenience function to generate the firm-wide report."""
    return FirmWideECReporter(
        aggregated_results=aggregated_results,
        copula_df=copula_df,
        output_dir=output_dir,
    ).generate_report()
